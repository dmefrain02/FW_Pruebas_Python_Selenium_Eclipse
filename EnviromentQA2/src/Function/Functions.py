#Libreria Selenium
import selenium
#Libreria Webdriver
from selenium import webdriver

#Librerias Webdrivers Services de los navegadores
from selenium.webdriver.chrome.service import Service as ChromeService 
from selenium.webdriver.firefox.service import Service as FirefoxService
from selenium.webdriver.edge.service import Service as EdgeService
#Librerias Webdrivers options de los navegadores
from selenium.webdriver.chrome.options import Options as OpcionesChrome
from selenium.webdriver.firefox.options import Options as OpcionesFirefox
from selenium.webdriver.edge.options import Options as OpcionesEdge

#Librerias Webdrivers Manager de los navegadores
from webdriver_manager.chrome import ChromeDriverManager
from webdriver_manager.firefox import GeckoDriverManager
from webdriver_manager.microsoft import EdgeChromiumDriverManager

#Librerias Webdrivers funcionalidades
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from selenium.webdriver.remote.webelement import isDisplayed_js
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys

from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.alert import Alert

from selenium.webdriver.remote import remote_connection

#from selenium.webdriver.common.print_page_options import PrintOptions as PrintOptions
#import base64
#import aspose.pdf as ap
#import urllib
from urllib3.exceptions import NewConnectionError, MaxRetryError

from Function.Inicializar import Inicializar
from Function.DriverFactory import DriverFactory
from selenium.common.exceptions import NoSuchElementException,NoAlertPresentException,NoSuchWindowException,TimeoutException, UnexpectedAlertPresentException, WebDriverException
import json
import pytest
from _ctypes_test import func
from locale import str
from _testmultiphase import Str
import time
import openpyxl
from idlelib.calltip import get_entity
import re # para expresiones regulares
import os # para capturas
import allure
import pyodbc
from allure_commons.types import AttachmentType
from pickle import NONE
from _pytest.python import Function
#from pkg_resources._vendor.jaraco.functools import except_
from _pytest.threadexception import catch_threading_exception
from http.cookiejar import Cookie
from PIL import Image #Pillow - Manejo de Imagenes
from io import BytesIO #Para conocer tamaños en bytes, ya esta instalado en Python

import pyautogui
import cv2
import numpy as np
from unittest.case import skip
from threading import Thread,Barrier
from numpy._core._multiarray_umath import rindex
from selenium.webdriver.common import desired_capabilities
from numpy.lib.tests.test__datasource import valid_baseurl
import subprocess
from sys import stderr
from email.errors import MessageError

class Functions(Inicializar):
    
    Nav_utilizado_capturas = ""
    
    #Abrir Navegador
    def abrir_navegador(self,navegador=Inicializar.Navegador,Remote = False, URL_SeleniumGrid = Inicializar.URL_SeleniumGrid):
        print(u"Directorio Base:" + Inicializar.BaseDir)
        print("-------------------------------------------")
        print(navegador)
        print("-------------------------------------------")
        
        self.Nav_utilizado_capturas = navegador          
        if navegador ==("Edge"):
            #Metodo para crear el driver de la instancia del Navegador
            self.driver = Functions.get_driver(self,navegador,Remote,URL_SeleniumGrid)
            
        elif navegador == ("Chrome"):
            #Metodo para crear el driver de la instancia del Navegador
            self.driver = Functions.get_driver(self,navegador,Remote,URL_SeleniumGrid)
            
        elif navegador ==("Firefox"):
            #Metodo para crear el driver de la instancia del Navegador
            self.driver = Functions.get_driver(self,navegador,Remote,URL_SeleniumGrid)
            
        elif navegador == ("Chrome_Remote"):
            #Metodo para crear el driver de la instancia del Navegador
            self.driver = Functions.get_driver(self,navegador,Remote,URL_SeleniumGrid)
            
        elif navegador == ("Edge_Remote"):
            #Metodo para crear el driver de la instancia del Navegador
            self.driver = Functions.get_driver(self,navegador,Remote,URL_SeleniumGrid)
        elif navegador == ("Firefox_Remote"):
            #Metodo para crear el driver de la instancia del Navegador
            self.driver = Functions.get_driver(self,navegador,Remote,URL_SeleniumGrid)
        elif navegador == ("Chrome_Docker"):
            self.driver = Functions.get_driver(self,navegador,Remote,URL_SeleniumGrid)
        elif navegador == ("Edge_Docker"):
            self.driver = Functions.get_driver(self,navegador,Remote,URL_SeleniumGrid)
        elif navegador == ("Firefox_Docker"):
            self.driver = Functions.get_driver(self,navegador,Remote,URL_SeleniumGrid)  
        else:
             raise ValueError(f"Navegador {navegador} no se encuentra soportado.")
        
        return self.driver

    #Retorna el Driver de la instancia del navegador a utilizar en las pruebas.
    def get_driver(self,navegador, remote, URL_Sel_Grid):
        self.driver = Functions._create_driver(self, navegador, remote, URL_Sel_Grid)
        return self.driver

     #Crear y configurar el driver: local o remoto
    def _create_driver(self, browser, remote, grid_url):   
        if remote:  #Instancia remota en Selenium Grid
            if browser == "Chrome_Remote":
                self.driver = Functions._create_chrome_remote_driver(self,grid_url)
            elif browser == "Edge_Remote":
                self.driver = Functions._create_edge_remote_driver(self,grid_url)
            elif browser == "Firefox_Remote":
                self.driver = Functions._create_firefox_remote_driver(self,grid_url)
            elif browser == "Chrome_Docker":
                self.driver = Functions._create_chrome_sel_grid_docker_driver(self,grid_url)
            elif browser == "Edge_Docker":
                self.driver = Functions._create_edge_sel_grid_docker_driver(self,grid_url)
            elif browser == "Firefox_Docker":
                self.driver = Functions._create_firefox_sel_grid_docker_driver(self,grid_url)
                
            return self.driver
        else:  #Instancia Navegador Local
            if browser == "Chrome":
                self.driver = Functions._create_chrome_driver(self)
                return self.driver
            elif browser == "Firefox":
                driver = Functions._create_firefox_driver(self)
                return self.driver
            elif browser == "Edge":
                driver = Functions._create_edge_driver(self)
                return self.driver

    #Crea y configura el driver de Chrome usando webdriver-manager
    def _create_chrome_driver(self):
        try:
            options = OpcionesChrome()
            prefs = {
                 "profile.default_content_settings.popups": 0,
                 "download.default_directory": Inicializar.Ruta_Descarga,
                 "directory_upgrade":True ,
                 "download.prompt_for_download": False,#Para que el navegador no pregunte al descargar
                 #"plugins.always_open_pdf_externally": True}) # Para que el navegador no abra el PDF en una pestaña nueva
                 #"plugins.plugins_disabled" : ["Chrome PDF Viewer"]
            }
            options.add_experimental_option("prefs", prefs)
            options.add_argument('start-maximized')
            #options.add_argument("headless")
            options.add_argument("--disable-extensions")#Deshabilita extensiones innecesarias
            
            chrome_driver_path = ChromeDriverManager().install() #Usa webdriver-manager para obtener la última versión compatible
            self.driver = webdriver.Chrome(service=ChromeService(chrome_driver_path), options=options)
            return self.driver
        except WebDriverException:
                print(u'No se abrio la instancia del navegador Local')
                Functions.cerrar_driver_navegador(self)
        except NewConnectionError:
                print(u'Error: No se logro abrir la instancia del navegador remoto.')
                raise ValueError(f"Error: Instancia del Navegador Remoto no se logro abrir.")
                Functions.cerrar_driver_navegador(self)
        except MaxRetryError:
                print(u'Error: No se logro abrir la instancia del navegador remoto.')
                raise ValueError(f"Error: Instancia del Navegador Remoto no se logro abrir.")
                Functions.cerrar_driver_navegador(self)
    
    #Crea y configura el driver de Firefox usando webdriver-manager
    def _create_firefox_driver(self):
        try:
            options = OpcionesFirefox()
            options.add_argument('--window-size=1200,1200')# Maximiza la ventana
            self.driver = webdriver.Firefox(service = FirefoxService(GeckoDriverManager().install()),options=options) #Usa webdriver-manager para obtener la última versión compatible
            return self.driver
        except WebDriverException:
                print(u'No se abrio la instancia del navegador Local')
                Functions.cerrar_driver_navegador(self)
        except NewConnectionError:
                print(u'Error: No se logro abrir la instancia del navegador remoto.')
                raise ValueError(f"Error: Instancia del Navegador Remoto no se logro abrir.")
                Functions.cerrar_driver_navegador(self)
        except MaxRetryError:
                print(u'Error: No se logro abrir la instancia del navegador remoto.')
                raise ValueError(f"Error: Instancia del Navegador Remoto no se logro abrir.")
                Functions.cerrar_driver_navegador(self)
    
    #Crea y configura el driver de Edge de manera local usando webdriver-manager
    def _create_edge_driver(self):
        try:
            options = OpcionesEdge()
            self.driver = webdriver.Edge(service =EdgeService(EdgeChromiumDriverManager().install()),options=options)
            self.driver.maximize_window()
            return self.driver
        except WebDriverException:
                print(u'No se abrio la instancia del navegador Local')
                Functions.cerrar_driver_navegador(self)
        except NewConnectionError:
                print(u'Error: No se logro abrir la instancia del navegador remoto.')
                raise ValueError(f"Error: Instancia del Navegador Remoto no se logro abrir.")
                Functions.cerrar_driver_navegador(self)
        except MaxRetryError:
                print(u'Error: No se logro abrir la instancia del navegador remoto.')
                raise ValueError(f"Error: Instancia del Navegador Remoto no se logro abrir.")
                Functions.cerrar_driver_navegador(self)
    
    #Crea y configura el driver de Chrome Remote de Selenium Grid
    def _create_chrome_remote_driver(self, grid_url):
        try:
            options = OpcionesChrome()
            prefs = {
                 "profile.default_content_settings.popups": 0,
                 "download.default_directory": Inicializar.Ruta_Descarga,
                 "directory_upgrade":True 
            }
            options.add_experimental_option("prefs",prefs)
            options.add_argument('start-maximized')
            self.driver = webdriver.Remote(grid_url,options=options)
            return self.driver
        except NewConnectionError:
                print(u'Error: No se logro abrir la instancia del navegador remoto.')
                raise ValueError(f"Error: Instancia del Navegador Remoto no se logro abrir.")
                Functions.cerrar_driver_navegador(self)
        except MaxRetryError:
                print(u'Error: No se logro abrir la instancia del navegador remoto.')
                raise ValueError(f"Error: Instancia del Navegador Remoto no se logro abrir.")
                Functions.cerrar_driver_navegador(self)
    
    #Crea y configura el driver de Edge Remote de Selenium Grid
    def _create_edge_remote_driver(self, grid_url):
        try:
            options = OpcionesEdge();
            options.add_argument("start-maximized")
            options.add_argument("inprivate")
            #options.add_argument("headless")
        
            self.driver = webdriver.Remote(grid_url,options=options)
            return self.driver
        except NewConnectionError:
                print(u'Error: No se logro abrir la instancia del navegador remoto.')
                raise ValueError(f"Error: Instancia del Navegador Remoto no se logro abrir.")
                Functions.cerrar_driver_navegador(self)
        except MaxRetryError:
                print(u'Error: No se logro abrir la instancia del navegador remoto.')
                raise ValueError(f"Error: Instancia del Navegador Remoto no se logro abrir.")
                Functions.cerrar_driver_navegador(self)
    
    #Crea y configura el driver de Firefox Remote de Selenium Grid
    def _create_firefox_remote_driver(self, grid_url):
        try:
            options = OpcionesFirefox()
            options.add_argument('--window-size=1200,1200')# Maximiza la ventana
            #options.add_argument("inprivate")
            #options.add_argument("headless")
            
            self.driver = webdriver.Remote(grid_url,options=options)
            return self.driver
        except NewConnectionError:
                print(u'Error: No se logro abrir la instancia del navegador remoto.')
                raise ValueError(f"Error: Instancia del Navegador Remoto no se logro abrir.")
                Functions.cerrar_driver_navegador(self)
        except MaxRetryError:
                print(u'Error: No se logro abrir la instancia del navegador remoto.')
                raise ValueError(f"Error: Instancia del Navegador Remoto no se logro abrir.")
                Functions.cerrar_driver_navegador(self)

    def _create_chrome_sel_grid_docker_driver(self, grid_url):
        try:
            options = webdriver.ChromeOptions()
            self.driver = webdriver.Remote(
                command_executor= grid_url,
                options=options
            ) 
            return self.driver
        except NewConnectionError:
                print(u'Error: No se logro abrir la instancia del navegador remoto.')
                raise ValueError(f"Error: Instancia del Navegador Remoto no se logro abrir.")
                Functions.cerrar_driver_navegador(self)
        except MaxRetryError:
                print(u'Error: No se logro abrir la instancia del navegador remoto.')
                raise ValueError(f"Error: Instancia del Navegador Remoto no se logro abrir.")
                Functions.cerrar_driver_navegador(self)
            
    def _create_edge_sel_grid_docker_driver(self, grid_url):
        try:
            options = OpcionesEdge();
            options.add_argument("start-maximized")
            options.add_argument("inprivate")
            #options.add_argument("headless")
            self.driver = webdriver.Remote(grid_url,options=options)
            return self.driver
        except NewConnectionError:
                print(u'Error: No se logro abrir la instancia del navegador remoto.')
                raise ValueError(f"Error: Instancia del Navegador Remoto no se logro abrir.")
                Functions.cerrar_driver_navegador(self)
        except MaxRetryError:
                print(u'Error: No se logro abrir la instancia del navegador remoto.')
                raise ValueError(f"Error: Instancia del Navegador Remoto no se logro abrir.")
                Functions.cerrar_driver_navegador(self)
            
    def _create_firefox_sel_grid_docker_driver(self, grid_url):
        try:
            options = OpcionesFirefox()
            options.add_argument('--window-size=1200,1200')# Maximiza la ventana
            #options.add_argument("inprivate")
            #options.add_argument("headless")
            self.driver = webdriver.Remote(grid_url,options=options)
            return self.driver
        except NewConnectionError:
                print(u'Error: No se logro abrir la instancia del navegador remoto.')
                raise ValueError(f"Error: Instancia del Navegador Remoto no se logro abrir.")
                Functions.cerrar_driver_navegador(self)
        except MaxRetryError:
                print(u'Error: No se logro abrir la instancia del navegador remoto.')
                raise ValueError(f"Error: Instancia del Navegador Remoto no se logro abrir.")
                Functions.cerrar_driver_navegador(self)

    #Cerrar la instancia del navegador
    def cerrar_driver_navegador(self):
        if self.driver:
            self.driver.quit()
        else:
            print("No hay un driver activo para cerrar.")

    #Dirigir a la URL del sitio de pruebas  
    def get_url_driver(self,URL):
        return self.driver.get(URL)
    
    #Metodo para encontrar elementos en el DOM
    def encontrando_elementos_en_el_DOM(self,estrategia_busqueda, valor_busqueda):
        try:
            if estrategia_busqueda == "xpath":
                elemento = self.driver.find_element(By.XPATH, valor_busqueda)
                print(u'ID_Elements: Se esta interactuando con el elemento: ' + valor_busqueda)
            elif estrategia_busqueda == "name":
                elemento = self.driver.find_element(By.NAME, valor_busqueda)
                print(u'ID_Elements: Se esta interactuando con el elemento: ' + valor_busqueda)
            elif estrategia_busqueda == "id":
                elemento = self.driver.find_element(By.ID, valor_busqueda)
                print(u'ID_Elements: Se esta interactuando con el elemento: ' + valor_busqueda)
            else:
                raise ValueError(f"Estrategia de busqueda {estrategia_busqueda} con el valor {valor_busqueda} no se encuentra soportada.")
            
            return elemento
            
        except TimeoutException:
            print(u'El elemento esperado no se presento: ' + valor_busqueda)
            Functions.cerrar_driver_navegador(self)
            
        except NoSuchElementException:
            print(u'El elemento esperado no se encuentro: ' + valor_busqueda)
            Functions.cerrar_driver_navegador(self)
    
    #Metodo para encontrar elementos en el DOM
    def encontrando_elementos_en_el_DOM_con_tiempo_espera(self,estrategia_busqueda, valor_busqueda, tiempo_espera):
        try:
            if estrategia_busqueda == "xpath":
                wait = WebDriverWait(self.driver,tiempo_espera)
                print(u'Esperando que el elemento se visualice: ' + valor_busqueda)
                wait.until(EC.visibility_of_element_located((By.XPATH,valor_busqueda)))
                wait.until(EC.element_to_be_clickable((By.XPATH,valor_busqueda)))
                                
                elemento = self.driver.find_element(By.XPATH, valor_busqueda)
                print('uElemento Encontrado:' + valor_busqueda)
                print(u'ID_Elements: Se esta interactuando con el elemento: ' + valor_busqueda)
            
            elif estrategia_busqueda == "name":
                wait = WebDriverWait(self.driver,tiempo_espera)
                print(u'Esperando que el elemento se visualice: ' + valor_busqueda)
                wait.until(EC.visibility_of_element_located((By.NAME,valor_busqueda)))
                wait.until(EC.element_to_be_clickable((By.NAME,valor_busqueda)))
                
                elemento = self.driver.find_element(By.NAME, valor_busqueda)
                print('uElemento Encontrado:' + valor_busqueda)
                print(u'ID_Elements: Se esta interactuando con el elemento: ' + valor_busqueda)
            
            elif estrategia_busqueda == "id":
                wait = WebDriverWait(self.driver,tiempo_espera)
                print(u'Esperando que el elemento se visualice: ' + valor_busqueda)
                wait.until(EC.visibility_of_element_located((By.ID,valor_busqueda)))
                wait.until(EC.element_to_be_clickable((By.ID,valor_busqueda)))

                elemento = self.driver.find_element(By.ID, valor_busqueda)
                print('uElemento Encontrado:' + valor_busqueda)
                print(u'ID_Elements: Se esta interactuando con el elemento: ' + valor_busqueda)
            
            else:
                raise ValueError(f"Estrategia de busqueda {estrategia_busqueda} con el valor {valor_busqueda} no se encuentra soportada.")
            
            return elemento
        except TimeoutException:
            print(u'El elemento esperado no se presento: ' + valor_busqueda)
            Functions.cerrar_driver_navegador(self)
            
        except NoSuchElementException:
            print(u'El elemento esperado no se encuentro: ' + valor_busqueda)
            Functions.cerrar_driver_navegador(self)
        except UnexpectedAlertPresentException:
            print(u'No se presento la alerta: ' + valor_busqueda)
            Functions.cerrar_driver_navegador(self)
    
    #Obtener Archivo JSON con los localizadores por medio del nombre      
    def obtener_archivo_json(self,file):
        json_ruta = Inicializar.Json + "/"+file+'.json'
        
        try:
            with open(json_ruta,'r')as read_file:
                self.json_strings = json.loads(read_file.read())
                print(u"Obtener Archivo Json: " + json_ruta)
                print(self.json_strings)
                return self.json_strings
            
        except FileNotFoundError:
            self.json_strings =False
            pytest.skip(u'Obtener Archivo Json: No se encontro el archivo json' + file)
            Functions.cerrar_driver_navegador(self)
     
    #Obtener entidad de elemento en el archivo JSON       
    def obtener_entidad(self,entidad):
        if self.json_strings is False:
            print(u'Define el DOM de la prueba')
        else:
            try:
                self.json_ValueToFind =self.json_strings[entidad]["ValueToFind"]
                self.json_GetFieldBy = self.json_strings[entidad]["GetFieldBy"]
                print(u"Se encontro la entidad: " + entidad  + " con los valores: " + self.json_ValueToFind + " y " + self.json_GetFieldBy)
                return True
            
            except KeyError:
                pytest.skip(u"Obtener Entidad: no se encontro la Key a la cual se hace referencia" + entidad)
                Functions.cerrar_driver_navegador(self)
                return None 
            
    #Obtener elemento apartir de la entidad en el archivo JSON  
    def obtener_elemento(self,entidad):
        Obtener_Entidad = Functions.obtener_entidad(self, entidad)
        
        if Obtener_Entidad is None:
            print (u'No se encontro el valor de la entidad en el doc. Json')
        else:
            try:
                elemento = Functions.encontrando_elementos_en_el_DOM(self, self.json_GetFieldBy, self.json_ValueToFind)
                
                print(u'Obtener Elemento: se encontro el elemento, ' + self.json_ValueToFind)
                return elemento
            except NoSuchElementException:
                print(u'Obtener Elemento: no se encontro el elemento, ' + self.json_ValueToFind)
                Functions.cerrar_driver_navegador(self)
            except TimeoutException:
                print(u'Obtener Elemento: no se encontro el elemento, ' + self.json_ValueToFind)
                Functions.cerrar_driver_navegador(self)          
    
    #Obtener texto en elemento
    def obtener_Texto(self,entidad):
        
        Get_Entity = Functions.obtener_entidad(self, entidad)     
        
        if Get_Entity is None:
            print(u'No se encontro el valor en el doc. Json')
        else:
            try:
                elemento = Functions.encontrando_elementos_en_el_DOM(self, self.json_GetFieldBy, self.json_ValueToFind) 
                    
                print(u'Obtener texto, se encontro el elemento: ' + self.json_ValueToFind)
                print(u'Text Value:' + elemento.text)
                return elemento.text
            except NoSuchElementException:
                print(u'Obtener Texto, No se encontro el elemento' + self.json_ValueToFind)
                Functions.cerrar_driver_navegador()
                
            except TimeoutException:
                print(u'Obtener Texto, No se encontro el elemento' + self.json_ValueToFind)
                Functions.cerrar_driver_navegador()

    #Espera informal
    def esperar_elemento(self,tiempo_espera = Inicializar.Tiempo_Espera):
        print("Inicia Espera: " +str(tiempo_espera))
        try:
            totalWait = 0
            while(totalWait < tiempo_espera):
                time.sleep(1)
                totalWait = totalWait + 1
                print("Tiempo total actual de espera: " + str(totalWait))
        finally:
            print("Espera: Carga Finalizada")
    
    #Hacer click en elemento 
    def click_en_elemento(self,entidad):
        Get_Entity = Functions.obtener_entidad(self,entidad)
        
        if Get_Entity is None:
            print(u'No se encontro el valor de la entidad buscada en el archivo .Json')
        else:
            try:
                elemento = Functions.encontrando_elementos_en_el_DOM(self, self.json_GetFieldBy, self.json_ValueToFind)
                    
                print(u'Elemento click en: '+ self.json_ValueToFind)
                return elemento.click()
            except NoSuchElementException:
                print(u'Elemento click, no se encontro el elemento: ' + self.json_ValueToFind)
                Functions.cerrar_driver_navegador(self)
            except TimeoutException:
                print(u'Elemento click, no se encontro el elemento: ' + self.json_ValueToFind)
                Functions.cerrar_driver_navegador(self)
    
    #Crear libro excel
    def crear_libro_excel(self,celda1=None, celda2=None, celda3=None,  celda4=None, celda5=None):
        #crear el libro
        wb = openpyxl.Workbook()
        
        #seleccionar e imprimir la hoja activa
        hoja = wb.active
        print(f'Hoja activa: {hoja.title}')
        
        #renombrar e imprimir la hoja activa
        hoja.title = "DataTest"
        print(f'Hoja activa: {hoja.title}') 
        
        hoja =wb["DataTest"]  
        if (celda1 != None)and (celda2 != None) and (celda3 != None):
            hoja[celda1] = "Valor Tomado"
            hoja[celda2] = "Valor Escrito"
            hoja[celda3] = "Resultado"
        
        if (celda4 != None)and (celda5 != None):
            hoja1 = wb.create_sheet("Bitacora Pruebas")
            hoja1[celda4] = "Prueba"
            hoja1[celda5] = "Resultado" 
            
        RutaArchivo = Inicializar.Excel_Crear + "\Pruebas.xlsx"     
        print(wb.sheetnames)
        print(u'Se creó el archivo en la ruta: ' + RutaArchivo)
        wb.save(RutaArchivo)     
    
    #Leer celda de excel   
    def leer_celda(self,celda, hoja):
        wb = openpyxl.load_workbook(Inicializar.Excel_Leer_Escribir)
        sheet = wb[hoja]
        valor = sheet[celda].value
        print(u'--------------------------------------------------')
        print(u'El libro de excel utilizado es: ' + Inicializar.Excel_Leer_Escribir)
        print(f'El valor de la celda es: {valor}')
        print(u'--------------------------------------------------')  
        return valor
    
    #Escribir en celda de excel
    def escribir_celda(self,celda,hoja,valor):
        wb = openpyxl.load_workbook(Inicializar.Excel_Leer_Escribir)
        hoja = wb[hoja]
        hoja[celda]=valor
        wb.save(Inicializar.Excel_Leer_Escribir)
        print(u'--------------------------------------------------')
        print(u'El libro de excel utilizado es: ' + Inicializar.Excel_Leer_Escribir)
        print(u'Se escribio en la celda: '+ celda + ' el valor: ' + valor)
        print(u'--------------------------------------------------')
    
    #Escribir texto en elemento
    def escribir_texto(self,entidad,texto):
        Get_Entidad = Functions.obtener_entidad(self, entidad)
        
        if Get_Entidad is None:
            print(u'No se encontro la entidad buscada: ' + entidad)
        else:
            try:
                elemento = Functions.encontrando_elementos_en_el_DOM(self, self.json_GetFieldBy, self.json_ValueToFind)
                print(f'Escribir texto en:  {self.json_ValueToFind} con el texto: {texto}')
                return elemento.send_keys(texto)
              
            except NoSuchElementException:
                print(u'Escribir Texto, No se encontro el elemento en el cual escribir: ' + self.json_ValueToFind)
                Functions.cerrar_driver_navegador()
                
            except TimeoutException:
                print(u'Escribir Texto, No se encontro el elemento en el cual escribir: ' + self.json_ValueToFind)
                Functions.cerrar_driver_navegador()
    
    #Presionar teclas especficias en elementos
    def envio_teclas_especificas(self,elemento,key):
        
        try:
            if key.lower()=='enter':
                Functions.obtener_elemento(self,elemento).send_keys(Keys.ENTER)
            if key.lower()=='tab':
                Functions.obtener_elemento(self,elemento).send_keys(Keys.TAB)
            if key.lower()=='space':
                Functions.obtener_elemento(self,elemento).send_keys(Keys.SPACE)   
                
        except TimeoutException:  
            print(u'No se logro realizar la acción con la tecla indicada ' + key + " en el elemento indicado: " + elemento)
            Functions.cerrar_driver_navegador(self)     
    
    #Limpiar elemento
    def limpiar_elemento(self,entidad):
        
        Get_Entidad = Functions.obtener_entidad(self, entidad)
        
        if Get_Entidad == None:
            print(u'No se encontro el valor de la entidad buscada en el doc. Json: ' + entidad)
            
        else:
            try:
                elemento = Functions.encontrando_elementos_en_el_DOM(self, self.json_GetFieldBy, self.json_ValueToFind)
                    
                print(u'Se limpio el texto en el elemento: ' + entidad + 'con el valor: ' + self.json_ValueToFind)
                return elemento.clear()
            
            except NoSuchElementException:
                print(u'Limpiar, no se encontro el elemento a limpiar: ' + self.json_ValueToFind + ' ' + entidad)
                Functions.cerrar_driver_navegador(self)
                
            except TimeoutException:
                print(u'Limpiar, no se encontro el elemento a limpiar: ' + self.json_ValueToFind + ' ' + entidad)
                Functions.cerrar_driver_navegador(self)
    
    #Espera explicita de elemento            
    def espera_explicita_elemento(self, locator):         
        Get_Entidad = Functions.obtener_entidad(self,locator)
        if Get_Entidad is None:
            return print(u'No se encontro el valor de la entidad requerida en el doc. JSON')
        else:
            try:
                if self.json_GetFieldBy.lower() == 'id':
                    wait =WebDriverWait(self.driver,15)
                    wait.until(EC.visibility_of_element_located((By.ID,self.json_ValueToFind)))
                    wait.until(EC.element_to_be_clickable((By.ID,self.json_ValueToFind)))   
                    print(u'Espera explicita, se visualizo el elemento: ' + locator + ' con el valor: ' + self.json_ValueToFind)
                    return True
                if self.json_GetFieldBy.lower() == 'xpath':
                    wait =WebDriverWait(self.driver,15)
                    wait.until(EC.visibility_of_element_located((By.XPATH,self.json_ValueToFind)))
                    wait.until(EC.element_to_be_clickable((By.XPATH,self.json_ValueToFind)))   
                    print(u'Espera explicita, se visualizo el elemento: ' + locator + ' con el valor: ' + self.json_ValueToFind)
                    return True
                if self.json_GetFieldBy.lower() == 'name':
                    wait =WebDriverWait(self.driver,15)
                    wait.until(EC.visibility_of_element_located((By.NAME,self.json_ValueToFind)))
                    wait.until(EC.element_to_be_clickable((By.NAME,self.json_ValueToFind)))   
                    print(u'Espera explicita, se visualizo el elemento: ' + locator + ' con el valor: ' + self.json_ValueToFind)
                    return True         
            except NoSuchElementException:
                print(u'Esperar explicita, no se encontro o no se visualizo el elemento luego de la espera: ' + locator + ' con el valor: ' + self.json_ValueToFind)
                Functions.cerrar_driver_navegador()
            except TimeoutException:
                print(u'Esperar explicita, no se encontro o no se visualizo el elemento luego de la espera: ' + locator + ' con el valor: ' + self.json_ValueToFind)
                Functions.cerrar_driver_navegador()             
    def WebdriverWait(self,time):
        WebDriverWait(self.driver,time)
    
    #Obtener fecha actual
    def obtener_fecha_actual(self):
        self.fecha = time.strftime(Inicializar.DateFormat)#Formato Fecha
        return self.fecha
    
    #Obtener hora actual
    def obtener_hora_actual(self):
        self.hora = time.strftime(Inicializar.HourFormat)#Formato 24Hrs
        return self.hora
    
    #Crear ruta para capturas de pantallas
    def crear_path(self):
        fecha = Functions.obtener_fecha_actual(self)
        GeneralPath = Inicializar.Path_Evidencias
        DriverTest = self.Nav_utilizado_capturas
            
        TestCase =self.__class__.__name__
        HoraActual = Functions.obtener_hora_actual(self)
        
        if   ((Inicializar.TestCase_x_Context =="S") and (GeneralPath != "")):
            path = f"{GeneralPath}\{fecha}\Pruebas\{TestCase}\{DriverTest}\{HoraActual}"
        elif ((Inicializar.TestCase_x_Context == "N") and (GeneralPath != "")):
            path =f"{GeneralPath}\{fecha}\{TestCase}\{DriverTest}\{HoraActual}"
        elif (((Inicializar.TestCase_x_Context == "N") or (Inicializar.TestCase_x_Context == "S")) and (GeneralPath == "")):
            path = f'{Inicializar.BaseDir}\Capturas\{fecha}\{TestCase}\{DriverTest}\{HoraActual}'
            print(f'No se encuentra establecida la ruta para guardar la captura de pantalla, se guardara en la carpeta raiz del framework de pruebas.\nEn: {path}')
        elif (((Inicializar.TestCase_x_Context !="S") or (Inicializar.TestCase_x_Context !="N") or (Inicializar.TestCase_x_Context == "")) and (GeneralPath == "")): 
            path = ""
            print(f'No se logro crear el path para guardar la captura de pantalla. Variables de "TestCase_x_Context" y "GeneralPath" no se han configurado correctamente: Tienen el valor: {GeneralPath} y {Inicializar.TestCase_x_Context}')
        
        if (path != ""):
            if not os.path.exists(path):
                os.makedirs(path)
            
            return path
        else:
            return Inicializar.Warning_Capturas
            
    #Realizar captura de pantalla
    def capturar_pantalla(self):
        Path=Functions.crear_path(self)
        TestCase =self.__class__.__name__
        
        if Path != Inicializar.Warning_Capturas:
            img = f'{Path}\{TestCase}\
            ('+Functions.obtener_fecha_actual(self)+' - '+ Functions.obtener_hora_actual(self)+')'+'.png'
            
            print(f'Se realizo captura de pantalla de la prueba: {img}')
            return self.driver.get_screenshot_as_file(img)
        else:
            print("Warning: No se logro generar la captura de pantalla. No se encuentra configurada el Path y variable contexto.")
    
    #Realizar captura de pantalla en reporte Allure
    def captura_pantalla_allure(self,Descripcion):
        allure.attach(self.driver.get_screenshot_as_png(),Descripcion,allure.attachment_type.PNG)
    
    #Realizar conexion a BD     
    def pyodbc_conexionBD(self,Env):
            
            try:
                if Env == 'DEV':
                    conn = pyodbc.connect(Inicializar.Cadena_Conexion_Dev)
                elif Env == 'QA':
                    conn = pyodbc.connect(Inicializar.Cadena_Conexion_QA)
                elif Env == 'UAT':
                    conn = pyodbc.connect(Inicializar.Cadena_Conexion_UAT)
                elif Env == 'PROD':
                    conn = pyodbc.connect(Inicializar.Cadena_Conexion_Prod)
                else:
                    print('No se logro establecer la cadena de conexion con la base de datos.')
                    
                self.cursor = conn.cursor()
                print("Conexion Exitosa")
                return self.cursor
            except (pyodbc.OperationalError) as Error:
                self.cursor = None
                print("Conexion Fallida")
                pytest.skip("Error en la conexion a la BD: ", str(Error))
    
    #Realizar consulta a BD         
    def pyodbc_ConsultaBD(self,Env,consulta_query):
        self.cursor = Functions.pyodbc_conexionBD(self,Env)
        if self.cursor is not None:
            try:
                self.cursor.execute(consulta_query)
                self.Result = self.cursor.fetchall()
                for row in self.Result:
                    return row[0]
                print(row[0])
                
            except (pyodbc.Error) as Error:
                print('Error en la consulta: ', Error)
                
            finally:
                if(self.cursor):
                    self.cursor.close()
                    print('pyodbc: Se cerro la conexion con la BD')                 
    
    #Obtener elemento select apartir de la entidad del archivo JSON               
    def obtener_elemento_select(self,entidad):
        Get_Entidad = Functions.obtener_entidad(self, entidad)
        if Get_Entidad is None:
            print(u'No se encontro el elemento ' + entidad + ' en el JSON definido')
        else:
            try:
                if  self.json_GetFieldBy.lower()== "id":
                    select = Select(self.driver.find_element(By.ID,self.json_ValueToFind))
                    print(u"get elements: " + self.json_ValueToFind) 
                if self.json_GetFieldBy.lower() == "name":
                    select = Select(self.driver.find_element(By.NAME,self.json_ValueToFind))
                    print(u"get elements: " + self.json_ValueToFind)
                                   
                if self.json_GetFieldBy.lower() == "xpath":
                    select = Select(self.driver.find_element(By.XPATH,self.json_ValueToFind))
                    print(u"get elements: " + self.json_ValueToFind)
    
                if self.json_GetFieldBy.lower() == "link":
                    select = Select(self.driver.find_element(By.LINK_TEXT,self.json_ValueToFind))
                    print(u"get elements: " + self.json_ValueToFind)
                 
                return select
            
            except NoSuchElementException:
                print(u"Select_Element: No presente: " + self.json_ValueToFind)
                Functions.cerrar_driver_navegador(self) 
            except TimeoutException:
                print(u"Select_Element: No presente: " + self.json_ValueToFind)
                Functions.cerrar_driver_navegador(self)
    
    #Obtener elemento select texto apartir de la entidad del archivo JSON    
    def obtener_elemento_select_texto(self,entidad,texto):
        select = Functions.obtener_elemento_select(self, entidad)
        select.select_by_visible_text(texto)
    
    #Realizar SCROLL en aplicativo 
    def Realizar_Scroll_JS(self,entidad):
        Get_Entidad = Functions.obtener_entidad(self, entidad)
        if Get_Entidad is None:
            return print(u'No se encontro el elemento: ' + entidad + ' en el JSON definido')
        else:
            try: 
                if self.json_GetFieldBy.lower() == 'id':
                    entidad = self.driver.find_element(By.ID, self.json_ValueToFind)
                    self.driver.execute_script("arguments[0].scrollIntoView();", entidad)
                    print(u'Se hizo scroll_to hasta el elemento')
                    return True
                if self.json_GetFieldBy.lower() == 'name':
                    entidad = self.driver.find_element(By.NAME, self.json_ValueToFind)
                    self.driver.execute_script("arguments[0].scrollIntoView();", entidad)
                    print(u'Se hizo scroll_to hasta el elemento')
                    return True
                if self.json_GetFieldBy.lower() == 'xpath':
                    entidad = self.driver.find_element(By.XPATH, self.json_ValueToFind)
                    self.driver.execute_script("arguments[0].scrollIntoView();", entidad)
                    print(u'Se hizo scroll_to hasta el elemento')
                    return True
                if self.json_GetFieldBy.lower() == 'class':
                    entidad = self.driver.find_element(By.CLASS_NAME, self.json_ValueToFind)
                    self.driver.execute_script("arguments[0].scrollIntoView();", entidad)
                    print(u'Se hizo scroll_to hasta el elemento')
                    return True
            except TimeoutException:
                print(u'JS Scroll: No se logro realizar hacia el elemento')
                Functions.cerrar_driver_navegador(self)
    
    #Abrir nuevo tab en instancia del navegador
    def abrir_nuevo_tab(self,Name_Tab,Page_Tab=Inicializar.Page_Tab): 
        self.driver.execute_script(f'''window.open("{Page_Tab}","{Name_Tab}");''')
        #self.driver.switch_to.new_window("tab")
        print(f'''window.open("{Page_Tab}","{Name_Tab}");''')
    
    #Abrir nuevo tab en nueva ventana del navegador
    def abrir_nuevo_tab_en_nueva_ventana(self,URL,Name_Tab,Page_Tab = Inicializar.Page_Tab):
        self.driver.execute_script(f'''window.open("{URL}","{Page_Tab}","{Name_Tab}");''')
        print(f'''window.open("{URL}","{Page_Tab}","{Name_Tab}");''')
    
    #Espera explicita
    def espera_explicita(self,driver,time, page_state):
        WebDriverWait(driver,time).until(lambda driver: page_state == 'complete')
        assert page_state == 'complete','No se completo la carga del sitio'
    
    #Moverse o intercambiar entre tabs abiertos en el navegador
    def intercambio_tab(self,ventana):
        self.driver.switch_to.window(self.driver.window_handles[ventana])
    
    #Cargar Archivo
    def cargar_archivo(self,entidad,txt_ruta_archivo = Inicializar.Archivo_Cargar):
        Get_Entidad = Functions.obtener_entidad(self, entidad)
        if Get_Entidad is None:
            return print(u'No se encontro el valor de la entidad requerida en el doc. Json')
        else:
            try:
                if self.json_GetFieldBy.lower()=='id':
                    Functions.escribir_texto(self, entidad, txt_ruta_archivo)
                    print(u'Se cargo el archivo en el elemento: '+ entidad + ' con el valor: ' + self.json_ValueToFind)
                if self.json_GetFieldBy.lower()=='name':
                    Functions.escribir_texto(self, entidad, txt_ruta_archivo)
                    print(u'Se cargo el archivo en el elemento: '+ entidad + ' con el valor: ' + self.json_ValueToFind)
                if self.json_GetFieldBy.lower()=='xpath':
                    Functions.escribir_texto(self, entidad, txt_ruta_archivo)
                    print(u'Se cargo el archivo en el elemento: '+ entidad + ' con el valor: ' + self.json_ValueToFind)
                    
            except NoSuchElementException:
                print(u'No se pudo cargar el archivo en el elemento: ' + entidad + ' con el valor: ' + self.json_ValueToFind)
                Functions.cerrar_driver_navegador(self)
            except TimeoutException:
                print(u'No se pudo cargar el archivo en el elemento: ' + entidad + ' con el valor: ' + self.json_ValueToFind)
                Functions.cerrar_driver_navegador(self)   
    
    #Assert
    def Assert_Equal(self,elemento,texto_esperado,msj):
        return self.assertEqual(Functions.obtener_Texto(self, elemento),texto_esperado, msj)
    def Assert_True(self,elemento,texto,msj):
        return self.assertTrue(Functions.obtener_Texto(self, elemento)==texto, msj)
    def AssertFalse(self,elemento,texto,msj):
        return self.assertFalse(Functions.obtener_Texto(self, elemento)==texto, msj)
    def Assert_True_IsDisplayer(self,elemento, msj):
        elemento = Functions.obtener_elemento(self, elemento)
        return self.assertTrue(elemento.is_displayed()==True,msj)
    def Assert_In_Elemento(self,texto_contenido, texto_ingresado_alert):
        return self.assertIn(texto_contenido, f'{texto_ingresado_alert}')

                         
    #Mover Mouse en aplicativo web
    def Mover_Mouse_x_App_Web(self,entidad):
        action = ActionChains(self.driver)
        elemento = Functions.obtener_elemento(self, entidad)
        
        try:
            action.move_to_element(elemento).perform()
            print(f'Mover Mouse: se movio el mouse al elemento: ' + self.json_ValueToFind)
            return elemento
        
        except NoSuchElementException:
            print(u'Mover_mouse: no se encontro el elemento, ' + self.json_ValueToFind)
            Functions.cerrar_driver_navegador(self)
        except TimeoutException:
            print(u'Mover_mouse: no se encontro el elemento, ' + self.json_ValueToFind)
            Functions.cerrar_driver_navegador(self)  
    
    #Arrastrar y Soltar: Drag & Drop
    def Arrastrar_y_Soltar(self,elemento_drag, elemento_drop):
        action = ActionChains(self.driver)
        element_drag = Functions.obtener_elemento(self, elemento_drag) 
        elemento_drop = Functions.obtener_elemento(self,elemento_drop) 
        
        try:
            #Perform drap and drop
            action.drag_and_drop(element_drag, elemento_drop).perform()  
        except NoSuchElementException:
            print(u'Drap_and_Drop: no se logro realizar la accion entre los elementos, ' + self.json_ValueToFind)
            Functions.cerrar_driver_navegador(self)
        except TimeoutException:
            print(u'Drap_and_Drop: no se logro realizar la accion entre los elementos, ' + self.json_ValueToFind)
            Functions.cerrar_driver_navegador(self)
        
    def Double_Click(self,elemento):
        action =ActionChains(self.driver)
        element = Functions.obtener_elemento(self, elemento)
        
        try:
            #hacer doble click
            action.double_click(element).perform()
        except NoSuchElementException:
            print(u'Double Click: no se logro realizar la accion, ' + self.json_ValueToFind)
            Functions.cerrar_driver_navegador(self)
        except TimeoutException:
            print(u'Double Click: no se logro realizar la accion, ' + self.json_ValueToFind)
            Functions.cerrar_driver_navegador(self)
    
    #Dar Click Derecho      
    def Click_Derecho(self,elemento):
        action =ActionChains(self.driver)
        element = Functions.obtener_elemento(self, elemento)
        
        try:
            #Click Derecho
            action.context_click(element).perform()
        except NoSuchElementException:
            print(u'Click Derecho: no se logro realizar la accion, ' + self.json_ValueToFind)
            Functions.cerrar_driver_navegador(self)
        except TimeoutException:
            print(u'Click Derecho: no se logro realizar la accion, ' + self.json_ValueToFind)
            Functions.cerrar_driver_navegador(self)
    
    #Mover Mouse entre elementos en aplicativo web       
    def Mover_Mouse(self,elemento1, elemento2):
        action =ActionChains(self.driver)
        element1 = Functions.obtener_elemento(self, elemento1)
        element2 = Functions.obtener_elemento(self, elemento2)
        
        try:
            #Mover Cursor
            action.move_to_element(element1).move_to_element(element2).perform()
        except NoSuchElementException:
            print(u'Mover Cursor: no se logro realizar la accion, ' + self.json_ValueToFind)
            Functions.cerrar_driver_navegador(self)
        except TimeoutException:
            print(u'Mover Cursor: no se logro realizar la accion, ' + self.json_ValueToFind)
            Functions.cerrar_driver_navegador(self)
    
    #Descargar Archivo
    def download_file(self,elemento):
        Functions.esperar_elemento(self)
        Functions.click_en_elemento(self, elemento)
        Functions.esperar_elemento(self)
        contenido = os.listdir(Inicializar.Ruta_Descarga)
        print(contenido)
        Functions.Assert_True_IsTrueDownload(self, contenido, f'No se descargo el archivo {Inicializar.Archivo_Descargado} en la ruta {Inicializar.Ruta_Descarga}')
    
    #Validar la descarga del archivo     
    def Assert_True_IsTrueDownload(self,contenido, msj):
        return self.assertTrue(Inicializar.Archivo_Descargado in contenido,msj)
        Functions.esperar_elemento(self)           
    
    #Escribir en un archivo bitacora
    def write_file_txt(self,texto,archivo):
        bitacora_pruebas = open(archivo, 'a')
        bitacora_pruebas.write(f'\nPrueba: {Functions.obtener_fecha_actual(self)} - {Functions.obtener_hora_actual(self)} - {texto} \n')
        bitacora_pruebas.write(f'--------------------------------------------------------------------')
        bitacora_pruebas.close()
    
    #Obtener cookie x nombre
    def obtener_cookie_nombre(self,nombre_cookie):
        print('--------------- Cookie x Nombre ----------------')
        cookie = self.driver.get_cookie(nombre_cookie) 
        print(cookie)
    
    #Obtener todas las cookies  
    def obtener_todas_las_cookies(self):
        print('--------------- Todas las cookies ----------------')
        cookie = self.driver.get_cookies() 
        print(cookie)      
    
    #Eliminar cookies x nombre
    def eliminar_cookie_x_nombre(self,nombre_cookie):
        print('-------------- Eliminar Cookie x Nombre ------------')
        self.driver.delete_cookie(nombre_cookie)
    
    #Eliminar todas las cookies     
    def eliminar_todas_las_cookies(self):
        print('-------------- Eliminar todas las cookies ---------------')
        self.driver.delete_all_cookies()
     
    #Realizar captura y cortar imagen   
    def captura_de_imagen_cortada(self, elemento_imagen, Ruta_Img_Cortada = Inicializar.Imagenes_Cortadas):
        imagen_encontrada = elemento_imagen.location
        size = elemento_imagen.size
        
        print(size)
        print(imagen_encontrada)
        
        imagen_guardada = self.driver.get_screenshot_as_png()
        imagen_cortada = Image.open(BytesIO(imagen_guardada))
        left = imagen_encontrada['x']
        top = imagen_encontrada['y']
        right = imagen_encontrada['x'] + size['width']
        bottom = imagen_encontrada['y'] +size['height']
        
        print(left,top,right,bottom)
        
        imagen_cortada = imagen_cortada.crop((left,top,right,bottom))
        imagen_cortada.save(f'{Ruta_Img_Cortada}\imagen_cortada.png')
    
    #Seleccionar fechas DTimePicker Dinamico
    def Selects_Fechas_DTPickerDinamico(self,AvanzarMes,Meses_Avanzar,FechaIda,FechaVuelta,DiaIda,DiaVuelta):
      #Contador para avanzar en el calendario
      avanzar = 0  
      
      #Click en control fecha ida
      Functions.click_en_elemento(self, FechaIda)
      Functions.esperar_elemento(self)
      
      #Avanzar en los meses del calendario     
      while (avanzar < Meses_Avanzar):
          Functions.click_en_elemento(self,AvanzarMes)
          avanzar = avanzar+1
    
      Functions.click_en_elemento(self, DiaIda)
      Functions.esperar_elemento(self)
      avanzar = 0
      
      #Avanzar en los meses del calendario  
      while (avanzar < Meses_Avanzar):
          Functions.click_en_elemento(self,AvanzarMes)
          avanzar = avanzar+1
          
      #Click en control fecha vuelta
      Functions.click_en_elemento(self, FechaVuelta)
      Functions.esperar_elemento(self)
      
      #Seleccionar dia vuelta
      Functions.click_en_elemento(self, DiaVuelta)
      Functions.esperar_elemento(self)
      
    #Seleccionar fechas DTimePicker Dinamico
    def Select_Fechas_DTPickerDinamicoUnico(self,FechaIda,Dia): 
      
        #Click en control fecha ida
        Functions.click_en_elemento(self, FechaIda)
        Functions.esperar_elemento(self)
        
        Functions.click_en_elemento(self, Dia)
        Functions.esperar_elemento(self)
     
    #Metodos para grabar videos (screen record) en las pruebas.  
    
    """Verifica si FFmpeg está disponible"""
    def inicializar_video(self, framerate = 30, formato=Inicializar.Formato_Video, nombrearchivo_video = Inicializar.VideoPruebas, output_dir = Inicializar.Carpeta_Videos):
              
        self.formato = formato.lower()
        self.output_video = output_dir
        self.filename = f"{nombrearchivo_video}.{formato}"
        self.output_file = os.path.join(self.output_video,self.filename)
        self.framerate = framerate
        os.makedirs(self.output_video, exist_ok=True)
        
        self.ffmpeg_path = Inicializar.ffmpeg_path
        self.ffmpeg_process = None

        # Validar si el formato es soportado
        '''if formato not in Inicializar.Formatos_Soportados:
            raise ValueError(f"Formato '{self.formato}' no soportado. Formatos válidos: {list(self.SUPPORTED_FORMATS.keys())}")
        else:
            print(f'Formatos soportados validos')'''

        # Verificar si FFmpeg está instalado
        if not Functions.valida_ffmpeg(self):
            raise FileNotFoundError("FFmpeg no encontrado. Agrega FFmpeg al PATH o proporciona la ruta completa.")
        else:
            print(f'FFmpeg validado en {Inicializar.ffmpeg_path}')
        
    """Verifica si FFmpeg está disponible en el sistema"""    
    def valida_ffmpeg(self):
        
        try:
            subprocess.run([self.ffmpeg_path, "-version"], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL, check=True)
            return True
        except (subprocess.CalledProcessError, FileNotFoundError):
            return False
    
    """Inicia la grabación de pantalla con FFmpeg"""   
    def start_recording(self):
        try:
            '''comando = [
               Inicializar.ffmpeg_path,
                "-y",  # Sobrescribe el archivo si ya existe
                "-f", "gdigrab",  # Captura la pantalla en Windows
                "-framerate", "30",  # FPS
                "-i", "desktop",  # Captura toda la pantalla
                "-c:v", "libx264", "-preset", "ultrafast", "-crf", "25",  # Compresión rápida
                self.output_file  # Nombre del archivo de salida
                ]'''     
            '''codec_info = Inicializar.Formatos_Soportados[self.formato]
            Inicializar.ffmpeg_path = [
                self.ffmpeg_path,
                "-y",
                "-f", "gdigrab",
                "-framerate", "30",
                "-i", "desktop",
                "-c:v", codec_info["codec"]
            ] + codec_info["extra"] + [self.output_file]'' 
            codec_info = Inicializar.Formatos_Soportados[self.formato]
            Inicializar.ffmpeg_path = [
                self.ffmpeg_path,
                "-y",
                "-f", "gdigrab",
                "-framerate", "30",
                "-i", "desktop",
                "-c:v", 
                codec_info["codec"],
                codec_info["extra"],
                self.output_file
            ]'''
                
            self.process = subprocess.Popen([
            "ffmpeg", "-y", "-f", "gdigrab", "-framerate", str(self.framerate),
            "-i", "desktop", self.output_file
        ], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            
            print(f"Grabación iniciada... Guardando en: {self.output_file}")
        except Exception as e:
            print(f"Error al iniciar FFmpeg: {e}")
    
    """Detiene la grabación de pantalla"""    
    def stop_recording(self):
        self.process.terminate()
        print(f"Grabación finalizada. Video guardado en: {self.output_file}")
    
    '''def inicializar_video(self,height_size, width_size,fps,nombre_arh_video,Ruta_Grabacion=Inicializar.Ruta_Grabacion):
        self.screen_size = (height_size,width_size)
        self.fps = fps
        self.output_filename = Ruta_Grabacion + f'\{nombre_arh_video}'
        
        self.format = cv2.VideoWriter_fourcc(*"XVID")
        self.salida = cv2.VideoWriter(self.output_filename,self.format,self.fps,self.screen_size)
        print('Se inicializo la grabacion')
        return self.salida   
    def grabar(self,salida):
        self.frame = pyautogui.screenshot()
        self.frame = np.array(self.frame)
        
        self.frame = cv2.cvtColor(self.frame,cv2.COLOR_BGR2RGB)
        
        salida.write(self.frame)
        return salida   
    def terminar_grabacion(self,salida):
        salida.release()
        cv2.destroyAllWindows()
        print('Se finaliza la grabación del video')'''
   
    def alert_navegadores(self,tipo_alert,texto_esperado="", msj="",texto_contenido ="", elemento="",texto_ingresado=""):
        self.alert = Alert(self.driver)
        self.text_alert = self.alert.text
        #0= alert OK por defecto, 1= alert espera de 5 segundos, 2=  ok y cancel, 3=  ok y cancel, 4= texto ingresado    
        #Alert con boton OK
        if (tipo_alert == 0):
            print('Mensaje de Alert: ' + self.text_alert)
            self.alert.accept()
        #Alert con espera de tiempo
        elif(tipo_alert == 1):
            print('Mensaje de Alert: ' + self.text_alert)
            self.alert.accept()
        #Alert con botones OK y Cancel
        elif(tipo_alert == 2):
            print('Mensaje de Alert: ' + self.text_alert)
            self.alert.accept()
            Functions.Assert_Equal(self,elemento, texto_esperado, msj)
        #Alert con botones OK y Cancel
        elif(tipo_alert == 3):
            print('Mensaje de Alert: ' + self.text_alert)
            self.alert.dismiss()
            Functions.Assert_Equal(self,elemento, texto_esperado, msj)
        #Alert con ingreso de texto
        elif(tipo_alert == 4):
            print('Mensaje de alert: ' + self.text_alert)
            self.alert.send_keys(texto_ingresado)
            self.alert.accept()
            Functions.Assert_In_Elemento(self,texto_contenido, Functions.obtener_Texto(self, elemento))
    
    ''' grabar segunda opcion, revisar graba bien solo el primer frame'''
    def inicializar(self, framerate = 30, formato=Inicializar.Formato_Video, nombrearchivo_video = Inicializar.VideoPruebas, output_dir = Inicializar.Carpeta_Videos):
        self.formato = formato.lower()
        self.output_video = output_dir
        self.filename = f"{nombrearchivo_video}.{self.formato}"
        self.output_file = os.path.join(self.output_video,self.filename)
        self.framerate = framerate
        self.resolution = pyautogui.size()
        os.makedirs(self.output_video, exist_ok=True)
        fourcc = cv2.VideoWriter_fourcc(*("XVID" if self.formato == "avi" else "mp4v"))
        self.out = cv2.VideoWriter(self.output_file,fourcc,self.framerate,self.resolution)
        
    def start(self,duracion = 10):
        start_time = time.time()
        while time.time()- start_time < duracion:
            img = pyautogui.screenshot()
            frame = np.array(img)
            #frame = cv2.cvtColor(frame,cv2.Color_RGB2BGR)
            self.out.write(frame)
            #self.stop()
        
    def stop(self):
        self.out.release()
        print(f"Video guardado en: {self.output_file}")
        
        
        
        