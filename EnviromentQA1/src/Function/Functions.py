#Libreria Selenium
import selenium
#Libreria Webdriver
from selenium import webdriver

#Librerias Webdrivers Services de los navegadores
from selenium.webdriver.chrome.service import Service as ChromeService 
from selenium.webdriver.firefox.service import Service as FirefoxService
from selenium.webdriver.edge.service import Service as EdgeService
#Librerias Webdrivers options de los navegadores
from selenium.webdriver.chrome.options import Options as OpcionesChrome
from selenium.webdriver.firefox.options import Options as OpcionesFirefox
from selenium.webdriver.edge.options import Options as OpcionesEdge

#Librerias Webdrivers Manager de los navegadores
from webdriver_manager.chrome import ChromeDriverManager
from webdriver_manager.firefox import GeckoDriverManager
from webdriver_manager.microsoft import EdgeChromiumDriverManager

#Librerias Webdrivers funcionalidades
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from selenium.webdriver.remote.webelement import isDisplayed_js
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys 
from selenium.webdriver.support.wait import WebDriverWait

from selenium.webdriver.remote import remote_connection
from selenium.webdriver.common.print_page_options import PrintOptions as PrintOptions
import base64

from Function.Inicializar import Inicializar
from selenium.common.exceptions import NoSuchElementException,NoAlertPresentException,NoSuchWindowException,TimeoutException
import json
import pytest
from _ctypes_test import func
from locale import str
from _testmultiphase import Str
import time
import openpyxl
from idlelib.calltip import get_entity
import re # para expresiones regulares
import os # para capturas
import allure
import pyodbc
from allure_commons.types import AttachmentType
from pickle import NONE
from _pytest.python import Function
#from pkg_resources._vendor.jaraco.functools import except_
from _pytest.threadexception import catch_threading_exception
from http.cookiejar import Cookie
from PIL import Image #Pillow - Manejo de Imagenes
from io import BytesIO #Para conocer tamaños en bytes, ya esta instalado en Python

import pyautogui
import cv2
import numpy as np
from unittest.case import skip


class Functions(Inicializar):
    
    Nav_utilizado_capturas = ""
    #Abrir Navegador
    def abrir_navegador(self,navegador=Inicializar.Navegador,URL_SeleniumGrid = Inicializar.URL_SeleniumGrid):
        print(u"Directorio Base:" + Inicializar.BaseDir)
        print("-------------------------------------------")
        print(navegador)
        print("-------------------------------------------")
        
        self.Nav_utilizado_capturas = navegador
        if navegador ==("Edge"):
            #Implementacion del Webdriver Manager            
            self.driver = webdriver.Edge(service =EdgeService(EdgeChromiumDriverManager().install()))
            self.driver.maximize_window()
            
        elif navegador == ("Chrome"):
            #Implementacion del WebdriverManager
            options = OpcionesChrome()
            prefs = {
                 "profile.default_content_settings.popups": 0,
                 "download.default_directory": Inicializar.Ruta_Descarga,
                 "directory_upgrade":True 
            }
            options.add_experimental_option("prefs", prefs)
            options.add_argument('start-maximized')
            #options.add_argument("headless")
            
            self.driver = webdriver.Chrome(service =ChromeService(ChromeDriverManager().install()),options=options)         
            
        elif navegador ==("Firefox"):
            #Implementacion WebDriver Manager
            options = OpcionesFirefox()
            options.add_argument('--window-size=800,800')

            self.driver = webdriver.Firefox(service = FirefoxService(GeckoDriverManager().install()),options=options)
        elif navegador == ("Chrome_Remote"):
            options = OpcionesChrome()
            prefs = {
                 "profile.default_content_settings.popups": 0,
                 "download.default_directory": Inicializar.Ruta_Descarga,
                 "directory_upgrade":True 
            }
            options.add_experimental_option("prefs",prefs)
            options.add_argument('start-maximized')
            self.driver = webdriver.Remote(URL_SeleniumGrid,options=options)
        elif navegador == ("Edge_Remote"):
            options = OpcionesEdge();
            options.add_argument("start-maximized")
            options.add_argument("inprivate")
            #options.add_argument("headless")
            self.driver = webdriver.Remote(URL_SeleniumGrid,options=options)
        
        return self.driver

    #Dirigir a la URL del sitio de pruebas  
    def get_url_driver(self,URL):
        return self.driver.get(URL)
    
    #Cerrar la instancia del navegador
    def cerrar_driver_navegador(self):
        return self.driver.quit()
    
    #Encontrando elementos en el DOM por medio de XPATH
    def elementos_del_DOM_x_XPATH(self, XPATH):
        elements = self.driver.find_element(By.XPATH, XPATH)
        print(u'XPATH_Elements: Se esta interactuando con el elemento: ' + XPATH)
        return elements
    
    #Encontrando elementos en el DOM por medio de ID
    def elementos_del_DOM_x_ID(self, ID):
        elements = self.driver.find_element(By.ID, ID)
        print(u'ID_Elements: Se esta interactuando con el elemento: ' + ID)
        return elements
    
    #Encontrando elementos en el DOM por medio de Name
    def elementos_del_DOM_x_NAME(self, NAME):
        elements = self.driver.find_element(By.NAME, NAME)
        print(u'NAME_Elements: Se esta interactuando con el elemento: ' + NAME)
        return elements
    
    #Encontrando elementos en el DOM por medio de XPATH utilizando esperas
    def _elementos_del_DOM_x_XPATH(self,XPATH):
        try:
            wait = WebDriverWait(self.driver,20)
            wait.until(EC.visibility_of_element_located((By.XPATH,XPATH)))
            wait.until(EC.element_to_be_clickable((By.XPATH,XPATH)))
            
            print(u'Esperando a que el elemento se visualice: ' + XPATH)
            elements = self.driver.find_element(By.XPATH, XPATH)
            
            print('uElemento Encontrado:' + XPATH)
            return elements
        except TimeoutException:
            print(u'Esperando el Elemento y este no se presento: ' + XPATH)
            Functions.cerrar_driver_navegador(self)
            
        except NoSuchElementException:
            print('uEsperando el Elemento y este no se presento: ' + XPATH)
            Functions.cerrar_driver_navegador(self)
    
    #Encontrando elementos en el DOM por medio de Id utilizando esperas    
    def _elementos_del_DOM_x_ID(self,ID):
        try:
            wait = WebDriverWait(self.driver,20)
            wait.until(EC.visibility_of_element_located((By.ID,ID)))
            wait.until(EC.element_to_be_clickable((By.ID,ID)))
            
            print(u'Esperando a que el elemento se visualice: ' + ID)
            elements = self.driver.find_element(By.ID, ID)
            
            print('uElemento Encontrado:' + ID)
            return elements
        except TimeoutException:
            print(u'Esperando el Elemento y este no se presento: ' + ID)
            Functions.cerrar_driver_navegador(self)
            
        except NoSuchElementException:
            print('uEsperando el Elemento y este no se presento: ' + ID)
            Functions.cerrar_driver_navegador(self)
    
    #Encontrando elementos en el DOM por medio de Name utilizando esperas        
    def _elementos_del_DOM_x_NAME(self,NAME):
        try:
            wait = WebDriverWait(self.driver,20)
            wait.until(EC.visibility_of_element_located((By.NAME,NAME)))
            wait.until(EC.element_to_be_clickable((By.NAME,NAME)))
            
            print(u'Esperando a que el elemento se visualice: ' + NAME)
            elements = self.driver.find_element(By.NAME, NAME)
            
            print('uElemento Encontrado:' + NAME)
            return elements
        except TimeoutException:
            print(u'Esperando el Elemento y este no se presento: ' + NAME)
            Functions.cerrar_driver_navegador(self)
            
        except NoSuchElementException:
            print('uEsperando el Elemento y este no se presento: ' + NAME)
            Functions.cerrar_driver_navegador(self)
    
    #Obtener Archivo JSON con los localizadores por medio del nombre      
    def obtener_archivo_json(self,file):
        json_ruta = Inicializar.Json + "/"+file+'.json'
        
        try:
            with open(json_ruta,'r')as read_file:
                self.json_strings = json.loads(read_file.read())
                print(u"Obtener Archivo Json: " + json_ruta)
                print(self.json_strings)
                return self.json_strings
            
        except FileNotFoundError:
            self.json_strings =False
            pytest.skip(u'Obtener Archivo Json: No se encontro el archivo json' + file)
            Functions.cerrar_driver_navegador(self)
     
    #Obtener entidad de elemento en el archivo JSON       
    def obtener_entidad(self,entidad):
        if self.json_strings is False:
            print(u'Define el DOM de la prueba')
        else:
            try:
                self.json_ValueToFind =self.json_strings[entidad]["ValueToFind"]
                self.json_GetFieldBy = self.json_strings[entidad]["GetFieldBy"]
                print(u"Se encontro la entidad: " + entidad  + " con los valores: " + self.json_ValueToFind + " y " + self.json_GetFieldBy)
                return True
            
            except KeyError:
                pytest.skip(u"Obtener Entidad: no se encontro la Key a la cual se hace referencia" + entidad)
                Functions.cerrar_driver_navegador(self)
                return None 
            
    #Obtener elemento apartir de la entidad en el archivo JSON  
    def obtener_elemento(self,entidad):
        Obtener_Entidad = Functions.obtener_entidad(self, entidad)
        
        if Obtener_Entidad is None:
            print (u'No se encontro el valor de la entidad en el doc. Json')
        else:
            try:
                if self.json_GetFieldBy.lower() == "id":
                    elemento = self.driver.find_element(By.ID, self.json_ValueToFind)
                if self.json_GetFieldBy.lower() == "name":
                    elemento = self.driver.find_element(By.NAME, self.json_ValueToFind)
                if self.json_GetFieldBy.lower() == "xpath":
                    elemento = self.driver.find_element(By.XPATH, self.json_ValueToFind)
                if self.json_GetFieldBy.lower() == "css":
                    elemento = self.driver.find_element(By.CSS_SELECTOR, self.json_ValueToFind)
                if self.json_GetFieldBy.lower() == "class":
                    elemento = self.driver.find_element(By.CLASS_NAME, self.json_ValueToFind)
                
                print(u'Obtener Elemento: se encontro el elemento, ' + self.json_ValueToFind)
                return elemento
            except NoSuchElementException:
                print(u'Obtener Elemento: no se encontro el elemento, ' + self.json_ValueToFind)
                Functions.cerrar_driver_navegador(self)
            except TimeoutException:
                print(u'Obtener Elemento: no se encontro el elemento, ' + self.json_ValueToFind)
                Functions.cerrar_driver_navegador(self)          
    
    #Obtener texto en elemento
    def obtener_Texto(self,entidad):
        
        Get_Entity = Functions.obtener_entidad(self, entidad)     
        
        if Get_Entity is None:
            print(u'No se encontro el valor en el doc. Json')
        else:
            try:
                if self.json_GetFieldBy.lower()=='id':
                    elemento = self.driver.find_element(By.ID,self.json_ValueToFind)
                if self.json_GetFieldBy.lower()=='name':
                    elemento = self.driver.find_element(By.NAME,self.json_ValueToFind)
                if self.json_GetFieldBy.lower()=='xpath':
                    elemento = self.driver.find_element(By.XPATH,self.json_ValueToFind)
                if self.json_GetFieldBy.lower()=='css':
                    elemento = self.driver.find_element(By.CSS_SELECTOR,self.json_ValueToFind)  
                    
                print(u'Obtener texto, se encontro el elemento: ' + self.json_ValueToFind)
                print(u'Text Value:' + elemento.text)
                return elemento.text
            except NoSuchElementException:
                print(u'Obtener Texto, No se encontro el elemento' + self.json_ValueToFind)
                Functions.cerrar_driver_navegador()
                
            except TimeoutException:
                print(u'Obtener Texto, No se encontro el elemento' + self.json_ValueToFind)
                Functions.cerrar_driver_navegador()

    #Espera informal
    def esperar_elemento(self,tiempo_espera = Inicializar.Tiempo_Espera):
        print("Inicia Espera: " +str(tiempo_espera))
        try:
            totalWait = 0
            while(totalWait < tiempo_espera):
                time.sleep(1)
                totalWait = totalWait + 1
                print("Tiempo total actual de espera: " + str(totalWait))
        finally:
            print("Espera: Carga Finalizada")
    
    #Hacer click en elemento 
    def click_en_elemento(self,entidad):
        Get_Entity = Functions.obtener_entidad(self,entidad)
        
        if Get_Entity is None:
            print(u'No se encontro el valor de la entidad buscada en el archivo .Json')
        else:
            try:
                if self.json_GetFieldBy.lower() == "id":
                    elemento = self.driver.find_element(By.ID, self.json_ValueToFind)
                if self.json_GetFieldBy.lower() == "xpath":
                    elemento = self.driver.find_element(By.XPATH, self.json_ValueToFind)
                if self.json_GetFieldBy.lower() == "name":
                    elemento = self.driver.find_element(By.NAME, self.json_ValueToFind)
                if self.json_GetFieldBy.lower() == "class":
                    elemento = self.driver.find_element(By.CLASS_NAME, self.json_ValueToFind)
                    
                print(u'Elemento click en: '+ self.json_ValueToFind)
                return elemento.click()
            except NoSuchElementException:
                print(u'Elemento click, no se encontro el elemento: ' + self.json_ValueToFind)
                Functions.cerrar_driver_navegador(self)
            except TimeoutException:
                print(u'Elemento click, no se encontro el elemento: ' + self.json_ValueToFind)
                Functions.cerrar_driver_navegador(self)
    
    #Crear libro excel
    def crear_libro_excel(self,celda1=None, celda2=None, celda3=None,  celda4=None, celda5=None):
        #crear el libro
        wb = openpyxl.Workbook()
        
        #seleccionar e imprimir la hoja activa
        hoja = wb.active
        print(f'Hoja activa: {hoja.title}')
        
        #renombrar e imprimir la hoja activa
        hoja.title = "DataTest"
        print(f'Hoja activa: {hoja.title}') 
        
        hoja =wb["DataTest"]  
        if (celda1 != None)and (celda2 != None) and (celda3 != None):
            hoja[celda1] = "Valor Tomado"
            hoja[celda2] = "Valor Escrito"
            hoja[celda3] = "Resultado"
        
        if (celda4 != None)and (celda5 != None):
            hoja1 = wb.create_sheet("Bitacora Pruebas")
            hoja1[celda4] = "Prueba"
            hoja1[celda5] = "Resultado" 
            
        RutaArchivo = Inicializar.Excel_Crear + "\Pruebas.xlsx"     
        print(wb.sheetnames)
        print(u'Se creó el archivo en la ruta: ' + RutaArchivo)
        wb.save(RutaArchivo)     
    
    #Leer celda de excel   
    def leer_celda(self,celda, hoja):
        wb = openpyxl.load_workbook(Inicializar.Excel_Leer_Escribir)
        sheet = wb[hoja]
        valor = sheet[celda].value
        print(u'--------------------------------------------------')
        print(u'El libro de excel utilizado es: ' + Inicializar.Excel_Leer_Escribir)
        print(f'El valor de la celda es: {valor}')
        print(u'--------------------------------------------------')  
        return valor
    
    #Escribir en celda de excel
    def escribir_celda(self,celda,hoja,valor):
        wb = openpyxl.load_workbook(Inicializar.Excel_Leer_Escribir)
        hoja = wb[hoja]
        hoja[celda]=valor
        wb.save(Inicializar.Excel_Leer_Escribir)
        print(u'--------------------------------------------------')
        print(u'El libro de excel utilizado es: ' + Inicializar.Excel_Leer_Escribir)
        print(u'Se escribio en la celda: '+ celda + ' el valor: ' + valor)
        print(u'--------------------------------------------------')
    
    #Escribir texto en elemento
    def escribir_texto(self,entidad,texto):
        Get_Entidad = Functions.obtener_entidad(self, entidad)
        
        if Get_Entidad is None:
            print(u'No se encontro la entidad buscada: ' + entidad)
        else:
            try:
                  if self.json_GetFieldBy.lower() == "id":
                     elemento = self.driver.find_element(By.ID,self.json_ValueToFind)
                  if self.json_GetFieldBy.lower() == "xpath":
                      elemento = self.driver.find_element(By.XPATH, self.json_ValueToFind)
                  if self.json_GetFieldBy.lower() == "name":
                     elemento = self.driver.find_element(By.NAME,self.json_ValueToFind)
                  if self.json_GetFieldBy.lower() == "class":
                      elemento = self.driver.find_element(By.CLASS_NAME, self.json_ValueToFind)
              
                  print(f'Escribir texto en:  {self.json_ValueToFind} con el texto: {texto}')
                  return elemento.send_keys(texto)
              
            except NoSuchElementException:
                print(u'Escribir Texto, No se encontro el elemento en el cual escribir: ' + self.json_ValueToFind)
                Functions.cerrar_driver_navegador()
                
            except TimeoutException:
                print(u'Escribir Texto, No se encontro el elemento en el cual escribir: ' + self.json_ValueToFind)
                Functions.cerrar_driver_navegador()
    
    #Presionar teclas especficias en elementos
    def envio_teclas_especificas(self,elemento,key):
        
        try:
            if key.lower()=='enter':
                Functions.obtener_elemento(self,elemento).send_keys(Keys.ENTER)
            if key.lower()=='tab':
                Functions.obtener_elemento(self,elemento).send_keys(Keys.TAB)
            if key.lower()=='space':
                Functions.obtener_elemento(self,elemento).send_keys(Keys.SPACE)   
                
        except TimeoutException:  
            print(u'No se logro realizar la acción con la tecla indicada ' + key + " en el elemento indicado: " + elemento)
            Functions.cerrar_driver_navegador(self)     
    
    #Limpiar elemento
    def limpiar_elemento(self,entidad):
        
        Get_Entidad = Functions.obtener_entidad(self, entidad)
        
        if Get_Entidad == None:
            print(u'No se encontro el valor de la entidad buscada en el doc. Json: ' + entidad)
            
        else:
            try:
                if self.json_GetFieldBy.lower() == 'id':
                    elemento = self.driver.find_element(By.ID, self.json_ValueToFind)
                if self.json_GetFieldBy.lower() == 'name':
                    elemento = self.driver.find_element(By.NAME, self.json_ValueToFind)
                if self.json_GetFieldBy.lower() == 'class':
                    elemento = self.driver.find_element(By.CLASS_NAME, self.json_ValueToFind)
                if self.json_GetFieldBy.lower() == 'xpath':
                    elemento = self.driver.find_element(By.XPATH, self.json_ValueToFind)
                    
                print(u'Se limpio el texto en el elemento: ' + entidad + 'con el valor: ' + self.json_ValueToFind)
                return elemento.clear()
            
            except NoSuchElementException:
                print(u'Limpiar, no se encontro el elemento a limpiar: ' + self.json_ValueToFind + ' ' + entidad)
                Functions.cerrar_driver_navegador(self)
                
            except TimeoutException:
                print(u'Limpiar, no se encontro el elemento a limpiar: ' + self.json_ValueToFind + ' ' + entidad)
                Functions.cerrar_driver_navegador(self)
    
    #Espera explicita de elemento            
    def espera_explicita_elemento(self, locator):
         
        Get_Entidad = Functions.obtener_entidad(self,locator)
        if Get_Entidad is None:
             return print(u'No se encontro el valor de la entidad requerida en el doc. JSON')
        else:
            try:
                 if self.json_GetFieldBy.lower() == 'id':
                     wait =WebDriverWait(self.driver,15)
                     wait.until(EC.visibility_of_element_located((By.ID,self.json_ValueToFind)))
                     wait.until(EC.element_to_be_clickable((By.ID,self.json_ValueToFind)))   
                     print(u'Espera explicita, se visualizo el elemento: ' + locator + ' con el valor: ' + self.json_ValueToFind)
                     return True
                 if self.json_GetFieldBy.lower() == 'xpath':
                     wait =WebDriverWait(self.driver,15)
                     wait.until(EC.visibility_of_element_located((By.XPATH,self.json_ValueToFind)))
                     wait.until(EC.element_to_be_clickable((By.XPATH,self.json_ValueToFind)))   
                     print(u'Espera explicita, se visualizo el elemento: ' + locator + ' con el valor: ' + self.json_ValueToFind)
                     return True
                 if self.json_GetFieldBy.lower() == 'name':
                     wait =WebDriverWait(self.driver,15)
                     wait.until(EC.visibility_of_element_located((By.NAME,self.json_ValueToFind)))
                     wait.until(EC.element_to_be_clickable((By.NAME,self.json_ValueToFind)))   
                     print(u'Espera explicita, se visualizo el elemento: ' + locator + ' con el valor: ' + self.json_ValueToFind)
                     return True           
            except NoSuchElementException:
                print(u'Esperar explicita, no se encontro o no se visualizo el elemento luego de la espera: ' + locator + ' con el valor: ' + self.json_ValueToFind)
                Functions.cerrar_driver_navegador()
            except TimeoutException:
                print(u'Esperar explicita, no se encontro o no se visualizo el elemento luego de la espera: ' + locator + ' con el valor: ' + self.json_ValueToFind)
                Functions.cerrar_driver_navegador()
    
    #Obtener fecha actual
    def obtener_fecha_actual(self):
        self.fecha = time.strftime(Inicializar.DateFormat)#Formato Fecha
        return self.fecha
    
    #Obtener hora actual
    def obtener_hora_actual(self):
        self.hora = time.strftime(Inicializar.HourFormat)#Formato 24Hrs
        return self.hora
    
    #Crear ruta para capturas de pantallas
    def crear_path(self):
        fecha = Functions.obtener_fecha_actual(self)
        GeneralPath = Inicializar.Path_Evidencias
        DriverTest = self.Nav_utilizado_capturas
            
        TestCase =self.__class__.__name__
        HoraActual = Functions.obtener_hora_actual(self)
        
        if   ((Inicializar.TestCase_x_Context =="S") and (GeneralPath != "")):
            path = f"{GeneralPath}\{fecha}\Pruebas\{TestCase}\{DriverTest}\{HoraActual}"
        elif ((Inicializar.TestCase_x_Context == "N") and (GeneralPath != "")):
            path =f"{GeneralPath}\{fecha}\{TestCase}\{DriverTest}\{HoraActual}"
        elif (((Inicializar.TestCase_x_Context == "N") or (Inicializar.TestCase_x_Context == "S")) and (GeneralPath == "")):
            path = f'{Inicializar.BaseDir}\Capturas\{fecha}\{TestCase}\{DriverTest}\{HoraActual}'
            print(f'No se encuentra establecida la ruta para guardar la captura de pantalla, se guardara en la carpeta raiz del framework de pruebas.\nEn: {path}')
        elif (((Inicializar.TestCase_x_Context !="S") or (Inicializar.TestCase_x_Context !="N") or (Inicializar.TestCase_x_Context == "")) and (GeneralPath == "")): 
            path = ""
            print(f'No se logro crear el path para guardar la captura de pantalla. Variables de "TestCase_x_Context" y "GeneralPath" no se han configurado correctamente: Tienen el valor: {GeneralPath} y {Inicializar.TestCase_x_Context}')
        
        if (path != ""):
            if not os.path.exists(path):
                os.makedirs(path)
            
            return path
        else:
            return "Error: Sin configurar el Path y el contexto para las capturas de pantalla."
                
    #Realizar captura de pantalla
    def capturar_pantalla(self):
        Path=Functions.crear_path(self)
        TestCase =self.__class__.__name__
        
        if Path != "Error: Sin configurar el Path y el contexto para las capturas de pantalla.":
            img = f'{Path}\{TestCase}\
            ('+Functions.obtener_fecha_actual(self)+' - '+ Functions.obtener_hora_actual(self)+')'+'.png'
            
            print(f'Se realizo captura de pantalla de la prueba: {img}')
            return self.driver.get_screenshot_as_file(img)
        else:
            print("Error: No se logro generar la captura de pantalla. No se encuentra configurada el Path y variable contexto.")
    
    #Realizar captura de pantalla en reporte Allure
    def captura_pantalla_allure(self,Descripcion):
        allure.attach(self.driver.get_screenshot_as_png(),Descripcion,allure.attachment_type.PNG)
    
    #Realizar conexion a BD     
    def pyodbc_conexionBD(self,Env):
            
            try:
                if Env == 'DEV':
                    conn = pyodbc.connect(Inicializar.Cadena_Conexion_Dev)
                elif Env == 'QA':
                    conn = pyodbc.connect(Inicializar.Cadena_Conexion_QA)
                elif Env == 'UAT':
                    conn = pyodbc.connect(Inicializar.Cadena_Conexion_UAT)
                elif Env == 'PROD':
                    conn = pyodbc.connect(Inicializar.Cadena_Conexion_Prod)
                else:
                    print('No se logro establecer la cadena de conexion con la base de datos.')
                    
                self.cursor = conn.cursor()
                print("Conexion Exitosa")
                return self.cursor
            except (pyodbc.OperationalError) as Error:
                self.cursor = None
                print("Conexion Fallida")
                pytest.skip("Error en la conexion a la BD: ", str(Error))
    
    #Realizar consulta a BD         
    def pyodbc_ConsultaBD(self,Env,consulta_query):
        self.cursor = Functions.pyodbc_conexionBD(self,Env)
        if self.cursor is not None:
            try:
                self.cursor.execute(consulta_query)
                self.Result = self.cursor.fetchall()
                for row in self.Result:
                    return row[0]
                print(row[0])
                
            except (pyodbc.Error) as Error:
                print('Error en la consulta: ', Error)
                
            finally:
                if(self.cursor):
                    self.cursor.close()
                    print('pyodbc: Se cerro la conexion con la BD')                 
    
    #Obtener elemento select apartir de la entidad del archivo JSON               
    def obtener_elemento_select(self,entidad):
        Get_Entidad = Functions.obtener_entidad(self, entidad)
        if Get_Entidad is None:
            print(u'No se encontro el elemento ' + entidad + ' en el JSON definido')
        else:
            try:
                if  self.json_GetFieldBy.lower()== "id":
                    select = Select(self.driver.find_element(By.ID,self.json_ValueToFind))
                    print(u"get elements: " + self.json_ValueToFind) 
                if self.json_GetFieldBy.lower() == "name":
                    select = Select(self.driver.find_element(By.NAME,self.json_ValueToFind))
                    print(u"get elements: " + self.json_ValueToFind)
                                   
                if self.json_GetFieldBy.lower() == "xpath":
                    select = Select(self.driver.find_element(By.XPATH,self.json_ValueToFind))
                    print(u"get elements: " + self.json_ValueToFind)
    
                if self.json_GetFieldBy.lower() == "link":
                    select = Select(self.driver.find_element(By.LINK_TEXT,self.json_ValueToFind))
                    print(u"get elements: " + self.json_ValueToFind)
                 
                return select
            
            except NoSuchElementException:
                print(u"Select_Element: No presente: " + self.json_ValueToFind)
                Functions.cerrar_driver_navegador(self) 
            except TimeoutException:
                print(u"Select_Element: No presente: " + self.json_ValueToFind)
                Functions.cerrar_driver_navegador(self)
    
    #Obtener elemento select texto apartir de la entidad del archivo JSON    
    def obtener_elemento_select_texto(self,entidad,texto):
        select = Functions.obtener_elemento_select(self, entidad)
        select.select_by_visible_text(texto)
    
    #Realizar SCROLL en aplicativo 
    def Realizar_Scroll_JS(self,entidad):
        Get_Entidad = Functions.obtener_entidad(self, entidad)
        if Get_Entidad is None:
            return print(u'No se encontro el elemento: ' + entidad + ' en el JSON definido')
        else:
            try: 
                if self.json_GetFieldBy.lower() == 'id':
                    entidad = self.driver.find_element(By.ID, self.json_ValueToFind)
                    self.driver.execute_script("arguments[0].scrollIntoView();", entidad)
                    print(u'Se hizo scroll_to hasta el elemento')
                    return True
                if self.json_GetFieldBy.lower() == 'name':
                    entidad = self.driver.find_element(By.NAME, self.json_ValueToFind)
                    self.driver.execute_script("arguments[0].scrollIntoView();", entidad)
                    print(u'Se hizo scroll_to hasta el elemento')
                    return True
                if self.json_GetFieldBy.lower() == 'xpath':
                    entidad = self.driver.find_element(By.XPATH, self.json_ValueToFind)
                    self.driver.execute_script("arguments[0].scrollIntoView();", entidad)
                    print(u'Se hizo scroll_to hasta el elemento')
                    return True
                if self.json_GetFieldBy.lower() == 'class':
                    entidad = self.driver.find_element(By.CLASS_NAME, self.json_ValueToFind)
                    self.driver.execute_script("arguments[0].scrollIntoView();", entidad)
                    print(u'Se hizo scroll_to hasta el elemento')
                    return True
            except TimeoutException:
                print(u'JS Scroll: No se logro realizar hacia el elemento')
                Functions.cerrar_driver_navegador(self)
    
    #Abrir nuevo tab en instancia del navegador
    def abrir_nuevo_tab(self,Name_Tab,Page_Tab=Inicializar.Page_Tab): 
        self.driver.execute_script(f'''window.open("{Page_Tab}","{Name_Tab}");''')
        #self.driver.switch_to.new_window("tab")
        print(f'''window.open("{Page_Tab}","{Name_Tab}");''')
    
    #Abrir nuevo tab en nueva ventana del navegador
    def abrir_nuevo_tab_en_nueva_ventana(self,URL,Name_Tab,Page_Tab = Inicializar.Page_Tab):
        self.driver.execute_script(f'''window.open("{URL}","{Page_Tab}","{Name_Tab}");''')
        print(f'''window.open("{URL}","{Page_Tab}","{Name_Tab}");''')
    
    #Espera explicita
    def espera_explicita(self,driver,time, page_state):
        WebDriverWait(driver,time).until(lambda driver: page_state == 'complete')
        assert page_state == 'complete','No se completo la carga del sitio'
    
    #Moverse o intercambiar entre tabs abiertos en el navegador
    def intercambio_tab(self,ventana):
        self.driver.switch_to.window(self.driver.window_handles[ventana])
    
    #Cargar Archivo
    def cargar_archivo(self,entidad,txt_ruta_archivo = Inicializar.Archivo_Cargar):
        Get_Entidad = Functions.obtener_entidad(self, entidad)
        if Get_Entidad is None:
            return print(u'No se encontro el valor de la entidad requerida en el doc. Json')
        else:
            try:
                if self.json_GetFieldBy.lower()=='id':
                    Functions.escribir_texto(self, entidad, txt_ruta_archivo)
                    print(u'Se cargo el archivo en el elemento: '+ entidad + ' con el valor: ' + self.json_ValueToFind)
                if self.json_GetFieldBy.lower()=='name':
                    Functions.escribir_texto(self, entidad, txt_ruta_archivo)
                    print(u'Se cargo el archivo en el elemento: '+ entidad + ' con el valor: ' + self.json_ValueToFind)
                if self.json_GetFieldBy.lower()=='xpath':
                    Functions.escribir_texto(self, entidad, txt_ruta_archivo)
                    print(u'Se cargo el archivo en el elemento: '+ entidad + ' con el valor: ' + self.json_ValueToFind)
                    
            except NoSuchElementException:
                print(u'No se pudo cargar el archivo en el elemento: ' + entidad + ' con el valor: ' + self.json_ValueToFind)
                Functions.cerrar_driver_navegador(self)
            except TimeoutException:
                print(u'No se pudo cargar el archivo en el elemento: ' + entidad + ' con el valor: ' + self.json_ValueToFind)
                Functions.cerrar_driver_navegador(self)   
    
    #Realizar Assert True para validar texto       
    def Assert_True_Validar_Texto(self,elemento,texto_esperado,msj):
        return self.assertTrue(Functions.obtener_Texto(self, elemento)==texto_esperado, msj)
        Functions.esperar_elemento(self)
        
     #Realizar Assert False para validar texto   
    def AssertFalse_Validar_Texto(self,elemento,texto_esperado,msj):
        return self.assertFalse(Functions.obtener_Texto(self, elemento)==texto_esperado, msj)
        Functions.esperar_elemento(self)
    
    #Realizar Assert Equals para validar texto
    def Assert_Equals_Comparar_Textos(self,elemento,texto_esperado,msj):
        return self.assertEqual(Functions.obtener_Texto(self, elemento),texto_esperado, msj)
        Functions.esperar_elemento(self) 
     
    #Realizar IsDisplayed para validar elemento
    def Assert_True_IsDisplayer_Elemento(self,elemento, msj):
        elemento = Functions.obtener_elemento(self, elemento)
        return self.assertTrue(elemento.is_displayed()==True,msj)
        Functions.esperar_elemento(self)
    
    #Mover Mouse en aplicativo web
    def Mover_Mouse_x_App_Web(self,entidad):
        action = ActionChains(self.driver)
        elemento = Functions.obtener_elemento(self, entidad)
        
        try:
            action.move_to_element(elemento).perform()
            print(f'Mover Mouse: se movio el mouse al elemento: ' + self.json_ValueToFind)
            return elemento
        
        except NoSuchElementException:
            print(u'Mover_mouse: no se encontro el elemento, ' + self.json_ValueToFind)
            Functions.cerrar_driver_navegador(self)
        except TimeoutException:
            print(u'Mover_mouse: no se encontro el elemento, ' + self.json_ValueToFind)
            Functions.cerrar_driver_navegador(self)  
    
    #Arrastrar y Soltar: Drag & Drop
    def Arrastrar_y_Soltar(self,elemento_drag, elemento_drop):
        action = ActionChains(self.driver)
        element_drag = Functions.obtener_elemento(self, elemento_drag) 
        elemento_drop = Functions.obtener_elemento(self,elemento_drop) 
        
        try:
            #Perform drap and drop
            action.drag_and_drop(element_drag, elemento_drop).perform()  
        except NoSuchElementException:
            print(u'Drap_and_Drop: no se logro realizar la accion entre los elementos, ' + self.json_ValueToFind)
            Functions.cerrar_driver_navegador(self)
        except TimeoutException:
            print(u'Drap_and_Drop: no se logro realizar la accion entre los elementos, ' + self.json_ValueToFind)
            Functions.cerrar_driver_navegador(self)
        
    def Double_Click(self,elemento):
        action =ActionChains(self.driver)
        element = Functions.obtener_elemento(self, elemento)
        
        try:
            #hacer doble click
            action.double_click(element).perform()
        except NoSuchElementException:
            print(u'Double Click: no se logro realizar la accion, ' + self.json_ValueToFind)
            Functions.cerrar_driver_navegador(self)
        except TimeoutException:
            print(u'Double Click: no se logro realizar la accion, ' + self.json_ValueToFind)
            Functions.cerrar_driver_navegador(self)
    
    #Dar Click Derecho      
    def Click_Derecho(self,elemento):
        action =ActionChains(self.driver)
        element = Functions.obtener_elemento(self, elemento)
        
        try:
            #Click Derecho
            action.context_click(element).perform()
        except NoSuchElementException:
            print(u'Click Derecho: no se logro realizar la accion, ' + self.json_ValueToFind)
            Functions.cerrar_driver_navegador(self)
        except TimeoutException:
            print(u'Click Derecho: no se logro realizar la accion, ' + self.json_ValueToFind)
            Functions.cerrar_driver_navegador(self)
    
    #Mover Mouse entre elementos en aplicativo web       
    def Mover_Mouse(self,elemento1, elemento2):
        action =ActionChains(self.driver)
        element1 = Functions.obtener_elemento(self, elemento1)
        element2 = Functions.obtener_elemento(self, elemento2)
        
        try:
            #Mover Cursor
            action.move_to_element(element1).move_to_element(element2).perform()
        except NoSuchElementException:
            print(u'Mover Cursor: no se logro realizar la accion, ' + self.json_ValueToFind)
            Functions.cerrar_driver_navegador(self)
        except TimeoutException:
            print(u'Mover Cursor: no se logro realizar la accion, ' + self.json_ValueToFind)
            Functions.cerrar_driver_navegador(self)
    
    #Descargar Archivo
    def download_file(self,elemento):
        Functions.esperar_elemento(self)
        Functions.click_en_elemento(self, elemento)
        Functions.esperar_elemento(self)
        contenido = os.listdir(Inicializar.Ruta_Descarga)
        print(contenido)
        Functions.Assert_True_IsTrueDownload(self, contenido, f'No se descargo el archivo {Inicializar.Archivo_Descargado} en la ruta {Inicializar.Ruta_Descarga}')
    
    #Validar la descarga del archivo     
    def Assert_True_IsTrueDownload(self,contenido, msj):
        return self.assertTrue(Inicializar.Archivo_Descargado in contenido,msj)
        Functions.esperar_elemento(self)           
    
    #Escribir en un archivo bitacora
    def write_file_txt(self,texto,archivo):
        bitacora_pruebas = open(archivo, 'a')
        bitacora_pruebas.write(f'\nPrueba: {Functions.obtener_fecha_actual(self)} - {Functions.obtener_hora_actual(self)} - {texto} \n')
        bitacora_pruebas.write(f'--------------------------------------------------------------------')
        bitacora_pruebas.close()
    
    #Obtener cookie x nombre
    def obtener_cookie_nombre(self,nombre_cookie):
        print('--------------- Cookie x Nombre ----------------')
        cookie = self.driver.get_cookie(nombre_cookie) 
        print(cookie)
    
    #Obtener todas las cookies  
    def obtener_todas_las_cookies(self):
        print('--------------- Todas las cookies ----------------')
        cookie = self.driver.get_cookies() 
        print(cookie)      
    
    #Eliminar cookies x nombre
    def eliminar_cookie_x_nombre(self,nombre_cookie):
        print('-------------- Eliminar Cookie x Nombre ------------')
        self.driver.delete_cookie(nombre_cookie)
    
    #Eliminar todas las cookies     
    def eliminar_todas_las_cookies(self):
        print('-------------- Eliminar todas las cookies ---------------')
        self.driver.delete_all_cookies()
     
    #Realizar captura y cortar imagen   
    def captura_de_imagen_cortada(self, elemento_imagen, Ruta_Img_Cortada = Inicializar.Imagenes_Cortadas):
        imagen_encontrada = elemento_imagen.location
        size = elemento_imagen.size
        
        print(size)
        print(imagen_encontrada)
        
        imagen_guardada = self.driver.get_screenshot_as_png()
        imagen_cortada = Image.open(BytesIO(imagen_guardada))
        left = imagen_encontrada['x']
        top = imagen_encontrada['y']
        right = imagen_encontrada['x'] + size['width']
        bottom = imagen_encontrada['y'] +size['height']
        
        print(left,top,right,bottom)
        
        imagen_cortada = imagen_cortada.crop((left,top,right,bottom))
        imagen_cortada.save(f'{Ruta_Img_Cortada}\imagen_cortada.png')
    
    #Seleccionar fechas DTimePicker Dinamico
    def Selects_Fechas_DTPickerDinamico(self,AvanzarMes,Meses_Avanzar,FechaIda,FechaVuelta,DiaIda,DiaVuelta):
      #Contador para avanzar en el calendario
      avanzar = 0  
      
      #Click en control fecha ida
      Functions.click_en_elemento(self, FechaIda)
      Functions.esperar_elemento(self)
      
      #Avanzar en los meses del calendario     
      while (avanzar < Meses_Avanzar):
          Functions.click_en_elemento(self,AvanzarMes)
          avanzar = avanzar+1
    
      Functions.click_en_elemento(self, DiaIda)
      Functions.esperar_elemento(self)
      avanzar = 0
      
      #Avanzar en los meses del calendario  
      while (avanzar < Meses_Avanzar):
          Functions.click_en_elemento(self,AvanzarMes)
          avanzar = avanzar+1
          
      #Click en control fecha vuelta
      Functions.click_en_elemento(self, FechaVuelta)
      Functions.esperar_elemento(self)
      
      #Seleccionar dia vuelta
      Functions.click_en_elemento(self, DiaVuelta)
      Functions.esperar_elemento(self)
      
    #Seleccionar fechas DTimePicker Dinamico
    def Select_Fechas_DTPickerDinamicoUnico(self,FechaIda,Dia): 
      
      #Click en control fecha ida
      Functions.click_en_elemento(self, FechaIda)
      Functions.esperar_elemento(self)
    
      Functions.click_en_elemento(self, Dia)
      Functions.esperar_elemento(self)
     
    #Metodos para grabar videos (screen record) en las pruebas.
    def inicializar_video(self,height_size, width_size,fps,nombre_arh_video,Ruta_Grabacion=Inicializar.Ruta_Grabacion):
        self.screen_size = (height_size,width_size)
        self.fps = fps
        self.output_filename = Ruta_Grabacion + f'\{nombre_arh_video}'
        
        self.format = cv2.VideoWriter_fourcc(*"XVID")
        self.salida = cv2.VideoWriter(self.output_filename,self.format,self.fps,self.screen_size)
        print('Se inicializo la grabacion')
        return self.salida   
    def grabar(self,salida):
        self.frame = pyautogui.screenshot()
        self.frame = np.array(self.frame)
        
        self.frame = cv2.cvtColor(self.frame,cv2.COLOR_BGR2RGB)
        
        salida.write(self.frame)
        return salida   
    def terminar_grabacion(self,salida):
        salida.release()
        cv2.destroyAllWindows()
        print('Se finaliza la grabación del video')
        
   # def print_page_pdf(self):
    #    pdf = self.driver.print_page()
     #   pdf.write(Path.get(f'{Inicializar.Path_Evidencias}/selenium.pdf'),OutputType.BYTES.convertFromBase64Png(pdf.getContent))